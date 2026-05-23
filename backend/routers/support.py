import os
import uuid
from datetime import datetime
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from jose import jwt, JWTError

from database import get_db
from models import SupportRequest, User
from schemas import SupportRequestCreate, SupportRequestResponse, ChatMessage, ChatResponse
from auth import oauth2_scheme, SECRET_KEY, ALGORITHM

router = APIRouter(prefix="/support", tags=["Support"])

EXCEL_FILE = "support_requests.xlsx"

# ─── In-memory chat session store ────────────────────────────────────────────
# { session_id: { "step": int, "data": { name, phone, category, description } } }
_sessions: Dict[str, Dict[str, Any]] = {}

STEPS = ["greet", "name", "phone", "category", "description", "confirm", "done"]
CATEGORIES = ["Points Issue", "Coupon Issue", "Store Issue", "App Issue", "Other"]


def get_optional_user(
    token: Optional[str] = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
) -> Optional[User]:
    """Retrieve current logged-in user optionally."""
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        phone_number: str = payload.get("sub")
        if phone_number:
            return db.query(User).filter(User.phone_number == phone_number).first()
    except JWTError:
        pass
    return None


def append_to_excel(request: SupportRequest):
    """Append a SupportRequest row to the Excel file."""
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Support Requests"
            ws.append([
                "Request ID", "User ID", "Name", "Phone", "Email",
                "Issue Category", "Description", "Created At (UTC)", "Status"
            ])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([
            request.id,
            request.user_id if request.user_id else "Guest",
            request.name,
            request.phone,
            request.email if request.email else "",
            request.issue_category,
            request.description,
            request.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            request.status
        ])
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False


def format_ticket_id(ticket_id: int) -> str:
    year = datetime.utcnow().year
    return f"SAV-{year}-{ticket_id:03d}"


def compute_wait_time() -> str:
    """Return a dynamic estimated wait time based on time of day (IST)."""
    hour = (datetime.utcnow().hour + 5) % 24  # rough IST offset
    if 9 <= hour < 13:
        return "~2 hours"
    elif 13 <= hour < 18:
        return "~4 hours"
    elif 18 <= hour < 21:
        return "~6 hours"
    else:
        return "~12 hours (next business day)"


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/contact")
def get_contact_info():
    """Return Savomart customer support contact details."""
    return {
        "phone": "1800-123-4567",
        "phone_hours": "Monday to Saturday, 9:00 AM – 7:00 PM IST",
        "email": "support@savomart.in",
        "email_response": "We respond within 24 hours",
        "whatsapp": "9876543210",
        "address": "Ebono Private Limited, Bangalore, Karnataka, India",
        "social": {
            "twitter": "@savomart",
            "instagram": "@savomart.in"
        }
    }


def _create_support_request(
    payload: SupportRequestCreate,
    current_user: Optional[User],
    db: Session
) -> dict:
    """Core logic for creating a support request — shared by both routes."""
    new_request = SupportRequest(
        user_id=current_user.id if current_user else None,
        name=payload.name,
        phone=payload.phone,
        email=payload.email,
        issue_category=payload.issue_category,
        description=payload.description,
        created_at=datetime.utcnow(),
        status="Open",
        saved_to_excel=False
    )

    db.add(new_request)
    db.commit()
    db.refresh(new_request)

    saved_ok = append_to_excel(new_request)
    if saved_ok:
        new_request.saved_to_excel = True
        db.commit()
        db.refresh(new_request)

    ticket_id = format_ticket_id(new_request.id)
    return {
        "ticket_id": ticket_id,
        "message": f"Your request has been received. Ticket ID: {ticket_id}. Our team will reach you within 24 hours.",
        "request": new_request,
    }


@router.post("", response_model=SupportRequestResponse)
def submit_support_request(
    payload: SupportRequestCreate,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Submit a loyalty support ticket. Saves to SQLite and Excel."""
    result = _create_support_request(payload, current_user, db)
    return result["request"]


@router.post("/request")
def submit_support_request_alias(
    payload: SupportRequestCreate,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Alias for submit_support_request — returns ticket_id and message."""
    result = _create_support_request(payload, current_user, db)
    return {
        "ticket_id": result["ticket_id"],
        "message": result["message"],
    }


@router.get("/my-tickets", response_model=List[SupportRequestResponse])
def get_user_support_tickets(
    current_user: User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Retrieve submitted support requests for the authenticated user."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required.")

    tickets = db.query(SupportRequest).filter(
        SupportRequest.user_id == current_user.id
    ).order_by(SupportRequest.created_at.desc()).all()

    return tickets


@router.post("/chat", response_model=ChatResponse)
def chat_with_savo(
    payload: ChatMessage,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Rule-based Savo AI assistant. Collects name/phone/category/description
    via a guided conversation, then auto-creates a support ticket.
    """
    session_id = payload.session_id
    user_msg = payload.message.strip()

    # Initialize session if new
    if session_id not in _sessions:
        _sessions[session_id] = {
            "step": "name",
            "data": {
                "name": current_user.name if current_user else None,
                "phone": current_user.phone_number if current_user else None,
                "category": None,
                "description": None,
            }
        }
        # Pre-fill from profile if logged in — skip name/phone steps
        session = _sessions[session_id]
        if current_user:
            session["step"] = "category"
            reply = (
                f"Hi {current_user.name}! 👋 I've pre-filled your details from your profile.\n\n"
                f"What's your issue about? Please pick a category:"
            )
            return ChatResponse(
                reply=reply,
                collected_data=session["data"],
                is_complete=False,
                quick_replies=CATEGORIES
            )
        else:
            reply = (
                "Hi there! 👋 I'm **Savo**, Savomart's AI support assistant. 🛒\n\n"
                "I'll help you raise a support request in just a few steps.\n\n"
                "First, could you tell me your **full name**?"
            )
            return ChatResponse(
                reply=reply,
                collected_data=session["data"],
                is_complete=False,
                quick_replies=[]
            )

    session = _sessions[session_id]
    step = session["step"]
    data = session["data"]

    # ── Step: collect name ────────────────────────────────────────────────────
    if step == "name":
        if len(user_msg) < 2:
            return ChatResponse(
                reply="Please enter your full name (at least 2 characters).",
                collected_data=data, is_complete=False, quick_replies=[]
            )
        data["name"] = user_msg.title()
        session["step"] = "phone"
        return ChatResponse(
            reply=f"Thanks **{data['name']}**! 😊 What's the best phone number to reach you at?",
            collected_data=data, is_complete=False, quick_replies=[]
        )

    # ── Step: collect phone ───────────────────────────────────────────────────
    if step == "phone":
        cleaned = "".join(filter(str.isdigit, user_msg))
        if len(cleaned) != 10:
            return ChatResponse(
                reply="Please enter a valid **10-digit Indian mobile number** (e.g. 9876543210).",
                collected_data=data, is_complete=False, quick_replies=[]
            )
        data["phone"] = cleaned
        session["step"] = "category"
        return ChatResponse(
            reply="Got it! What's your issue about? Please pick a category:",
            collected_data=data, is_complete=False, quick_replies=CATEGORIES
        )

    # ── Step: collect category ────────────────────────────────────────────────
    if step == "category":
        # Accept either a category chip click or typed text
        matched = next((c for c in CATEGORIES if c.lower() in user_msg.lower()), None)
        if not matched:
            return ChatResponse(
                reply="Please select one of the categories below:",
                collected_data=data, is_complete=False, quick_replies=CATEGORIES
            )
        data["category"] = matched
        session["step"] = "description"
        return ChatResponse(
            reply=f"Got it — **{matched}**. 📝\n\nPlease briefly describe your issue and I'll get it logged for you right away.",
            collected_data=data, is_complete=False, quick_replies=[]
        )

    # ── Step: collect description ─────────────────────────────────────────────
    if step == "description":
        if len(user_msg) < 10:
            return ChatResponse(
                reply="Could you give a bit more detail? (at least 10 characters)",
                collected_data=data, is_complete=False, quick_replies=[]
            )
        data["description"] = user_msg
        session["step"] = "confirm"
        summary = (
            f"Here's a summary of your request:\n\n"
            f"📋 **Name:** {data['name']}\n"
            f"📱 **Phone:** {data['phone']}\n"
            f"🏷️ **Category:** {data['category']}\n"
            f"📝 **Issue:** {data['description']}\n\n"
            f"Shall I submit this ticket?"
        )
        return ChatResponse(
            reply=summary,
            collected_data=data, is_complete=False, quick_replies=["Yes, submit!", "No, let me redo"]
        )

    # ── Step: confirm and create ticket ──────────────────────────────────────
    if step == "confirm":
        if "no" in user_msg.lower() or "redo" in user_msg.lower() or "cancel" in user_msg.lower():
            # Reset to beginning
            _sessions[session_id] = {
                "step": "name" if not current_user else "category",
                "data": {
                    "name": current_user.name if current_user else None,
                    "phone": current_user.phone_number if current_user else None,
                    "category": None,
                    "description": None,
                }
            }
            start_step = "category" if current_user else "name"
            first_q = (
                "No problem! Let's start over.\n\nWhat category does your issue fall under?"
                if current_user else
                "No problem! Let's start over.\n\nWhat's your full name?"
            )
            return ChatResponse(
                reply=first_q,
                collected_data=_sessions[session_id]["data"],
                is_complete=False,
                quick_replies=CATEGORIES if current_user else []
            )

        if "yes" in user_msg.lower() or "submit" in user_msg.lower():
            # Create the ticket
            new_request = SupportRequest(
                user_id=current_user.id if current_user else None,
                name=data["name"],
                phone=data["phone"],
                email=None,
                issue_category=data["category"],
                description=data["description"],
                created_at=datetime.utcnow(),
                status="Open",
                saved_to_excel=False
            )
            db.add(new_request)
            db.commit()
            db.refresh(new_request)

            saved_ok = append_to_excel(new_request)
            if saved_ok:
                new_request.saved_to_excel = True
                db.commit()
                db.refresh(new_request)

            ticket_id = format_ticket_id(new_request.id)
            session["step"] = "done"
            data["ticket_id"] = ticket_id

            return ChatResponse(
                reply=(
                    f"✅ **Done!** Your ticket has been raised.\n\n"
                    f"🎫 **Ticket ID:** `{ticket_id}`\n"
                    f"Our team will contact you within 24 hours at **{data['phone']}**.\n\n"
                    f"Is there anything else I can help with?"
                ),
                collected_data=data,
                is_complete=True,
                ticket_id=ticket_id,
                quick_replies=["No, I'm good!", "Yes, raise another"]
            )

        return ChatResponse(
            reply="Please reply **Yes** to submit or **No** to start over.",
            collected_data=data, is_complete=False, quick_replies=["Yes, submit!", "No, let me redo"]
        )

    # ── Step: done — handle follow-up ─────────────────────────────────────────
    if step == "done":
        if "another" in user_msg.lower() or "yes" in user_msg.lower():
            _sessions[session_id] = {
                "step": "category" if current_user else "name",
                "data": {
                    "name": current_user.name if current_user else None,
                    "phone": current_user.phone_number if current_user else None,
                    "category": None,
                    "description": None,
                }
            }
            return ChatResponse(
                reply="Sure! Let's raise another request.\n\nWhat category does your new issue fall under?",
                collected_data=_sessions[session_id]["data"],
                is_complete=False,
                quick_replies=CATEGORIES
            )
        return ChatResponse(
            reply="Happy to help! 😊 Feel free to come back anytime you need support. Take care!",
            collected_data=data, is_complete=True, quick_replies=[]
        )

    # Fallback
    return ChatResponse(
        reply="I didn't quite catch that. Could you try again?",
        collected_data=data, is_complete=False, quick_replies=[]
    )
