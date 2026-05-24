from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import SupportRequest, Admin
from passlib.context import CryptContext
from pydantic import BaseModel
from datetime import datetime
import os
import openpyxl

router = APIRouter()
pwd_context = CryptContext(
    schemes=["bcrypt"], deprecated="auto"
)

class AdminLoginSchema(BaseModel):
    email: str
    password: str

class StatusUpdateSchema(BaseModel):
    status: str
    resolution_note: str = ""
    resolved_by: str = "Admin"

@router.put("/tickets/{ticket_id}/status")
def update_status(
    ticket_id: int,
    payload: StatusUpdateSchema,
    db: Session = Depends(get_db)
):
    valid = ["Open", "InProgress", "Resolved"]
    if payload.status not in valid:
        raise HTTPException(
            status_code=400,
            detail=f"Must be one of {valid}"
        )
    t = db.query(SupportRequest).filter(
        SupportRequest.id == ticket_id
    ).first()
    if not t:
        raise HTTPException(
            status_code=404,
            detail="Ticket not found"
        )
    old_status = t.status
    t.status = payload.status
    # When marking as resolved — set timestamp and note
    if payload.status == "Resolved" and old_status != "Resolved":
        t.resolved_at = datetime.utcnow()
        t.resolved_by = payload.resolved_by or "Admin"
        t.resolution_note = payload.resolution_note or ""
        t.customer_feedback = None  # reset feedback
        print(f"Ticket {ticket_id} resolved at {t.resolved_at} by {t.resolved_by}")
    # When reopening — clear resolution data and increment count
    if payload.status == "Open" and old_status == "Resolved":
        t.resolved_at = None
        t.resolved_by = None
        t.resolution_note = None
        t.customer_feedback = None
        t.reopen_count = (t.reopen_count or 0) + 1
    db.commit()
    db.refresh(t)
    return {
        "message": "Status updated",
        "status": t.status,
        "resolved_at": t.resolved_at.isoformat() if t.resolved_at else None,
        "resolution_note": t.resolution_note,
        "resolved_by": t.resolved_by,
        "customer_feedback": t.customer_feedback,
        "reopen_count": t.reopen_count,
    }

class FeedbackSchema(BaseModel):
    feedback: str
    reason: str = ""

@router.post("/tickets/{ticket_id}/feedback")
def submit_customer_feedback(
    ticket_id: int,
    payload: FeedbackSchema,
    db: Session = Depends(get_db)
):
    t = db.query(SupportRequest).filter(
        SupportRequest.id == ticket_id
    ).first()
    if not t:
        raise HTTPException(
            status_code=404,
            detail="Ticket not found"
        )
    if t.status != "Resolved":
        raise HTTPException(
            status_code=400,
            detail="Can only give feedback on resolved tickets"
        )
    valid_feedback = ["Accepted", "Reopened"]
    if payload.feedback not in valid_feedback:
        raise HTTPException(
            status_code=400,
            detail=f"Feedback must be: {valid_feedback}"
        )
    t.customer_feedback = payload.feedback
    t.feedback_at = datetime.utcnow()
    if payload.feedback == "Reopened":
        t.status = "Open"
        t.reopen_reason = payload.reason or ""
        t.reopen_count = (t.reopen_count or 0) + 1
        t.resolved_at = None
        t.resolution_note = None
        t.resolved_by = None
        print(f"Ticket {ticket_id} reopened by customer")
    else:
        print(f"Ticket {ticket_id} accepted by customer")
    db.commit()
    return {
        "message": "Feedback submitted",
        "feedback": t.customer_feedback,
        "status": t.status,
        "feedback_at": t.feedback_at.isoformat() if t.feedback_at else None,
        "reopen_reason": t.reopen_reason,
        "reopen_count": t.reopen_count,
    }

def create_admin_token(admin_id: int, email: str):
    from jose import jwt
    import os
    SECRET = os.getenv("SECRET_KEY", "savomart-secret-2025")
    data = {
        "sub": str(admin_id),
        "email": email,
        "role": "admin",
        "exp": datetime.utcnow().timestamp() + 86400 * 7
    }
    return jwt.encode(data, SECRET, algorithm="HS256")

@router.post("/login")
def admin_login(
    payload: AdminLoginSchema,
    db: Session = Depends(get_db)
):
    print(f"Admin login attempt: {payload.email}")
    
    admin = db.query(Admin).filter(
        Admin.email == payload.email.strip().lower()
    ).first()
    
    if not admin:
        print(f"Admin not found: {payload.email}")
        raise HTTPException(
            status_code=401,
            detail="Invalid email or password"
        )
    
    if not pwd_context.verify(
        payload.password, admin.password_hash
    ):
        print("Password verification failed")
        raise HTTPException(
            status_code=401,
            detail="Invalid email or password"
        )
    
    token = create_admin_token(admin.id, admin.email)
    print(f"Admin login success: {admin.email}")
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "admin": {
            "id": admin.id,
            "name": admin.name,
            "email": admin.email
        }
    }

@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    print("Stats endpoint called")
    try:
        total = db.query(SupportRequest).count()
        open_c = db.query(SupportRequest).filter(
            SupportRequest.status == "Open"
        ).count()
        inp = db.query(SupportRequest).filter(
            SupportRequest.status == "InProgress"
        ).count()
        res = db.query(SupportRequest).filter(
            SupportRequest.status == "Resolved"
        ).count()
        
        today = datetime.utcnow().strftime("%Y-%m-%d")
        today_c = db.query(SupportRequest).filter(
            SupportRequest.created_at >= today
        ).count()
        
        cats = db.query(
            SupportRequest.issue_category,
            func.count(SupportRequest.id)
        ).group_by(
            SupportRequest.issue_category
        ).all()
        
        result = {
            "total": total,
            "open": open_c,
            "in_progress": inp,
            "resolved": res,
            "today": today_c,
            "categories": [
                {"category": c[0], "count": c[1]}
                for c in cats
            ]
        }
        print(f"Stats result: {result}")
        return result
        
    except Exception as e:
        print(f"Stats error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

@router.get("/tickets")
def get_tickets(
    status: str = Query(None),
    category: str = Query(None),
    search: str = Query(None),
    skip: int = Query(0),
    limit: int = Query(20),
    db: Session = Depends(get_db)
):
    print(f"Tickets called: status={status} "
          f"category={category} search={search}")
    try:
        q = db.query(SupportRequest)
        
        if status:
            q = q.filter(SupportRequest.status == status)
        if category:
            q = q.filter(
                SupportRequest.issue_category == category
            )
        if search:
            s = f"%{search}%"
            q = q.filter(
                SupportRequest.name.ilike(s) |
                SupportRequest.phone.ilike(s)
            )
        
        total = q.count()
        print(f"Total matching tickets: {total}")
        
        items = q.order_by(
            SupportRequest.created_at.desc()
        ).offset(skip).limit(limit).all()
        
        result = []
        for t in items:
            result.append({
                "id": t.id,
                "ticket_id": f"SAV-2025-{t.id:03d}",
                "name": t.name or "",
                "phone": t.phone or "",
                "email": t.email or "",
                "issue_category": t.issue_category or "",
                "description": t.description or "",
                "status": t.status or "Open",
                "created_at": (
                    t.created_at.isoformat()
                    if t.created_at
                    else datetime.utcnow().isoformat()
                )
            })
        
        return {
            "tickets": result,
            "total": total,
            "pages": max(1, (total + limit - 1) // limit)
        }
        
    except Exception as e:
        print(f"Tickets error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

@router.put("/tickets/{ticket_id}/status")
def update_status(
    ticket_id: int,
    payload: StatusUpdateSchema,
    db: Session = Depends(get_db)
):
    valid = ["Open", "InProgress", "Resolved"]
    if payload.status not in valid:
        raise HTTPException(
            status_code=400,
            detail=f"Must be one of {valid}"
        )
    
    t = db.query(SupportRequest).filter(
        SupportRequest.id == ticket_id
    ).first()
    
    if not t:
        raise HTTPException(
            status_code=404,
            detail="Ticket not found"
        )
    
    t.status = payload.status
    db.commit()
    return {
        "message": "Updated",
        "status": payload.status
    }

@router.get("/download-excel")
def download_excel(db: Session = Depends(get_db)):
    try:
        filepath = os.path.join(
            os.getcwd(), "support_requests.xlsx"
        )
        print(f"Generating Excel at: {filepath}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Support Tickets"
        
        # Style header
        headers = [
            "Ticket ID", "Name", "Phone", "Email",
            "Category", "Description", "Status",
            "Created At", "Resolved At", "Resolution Note",
            "Resolved By", "Customer Feedback", "Feedback Date",
            "Reopen Reason", "Times Reopened"
        ]
        ]
        ws.append(headers)
        
        # Style header row
        from openpyxl.styles import (
            PatternFill, Font, Alignment
        )
        header_fill = PatternFill(
            start_color="782B90",
            end_color="782B90",
            fill_type="solid"
        )
        header_font = Font(
            color="FFFFFF", bold=True
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add all tickets from DB
        tickets = db.query(SupportRequest).order_by(
            SupportRequest.created_at.desc()
        ).all()
        
        for t in tickets:
                    ws.append([
            f"SAV-2025-{t.id:03d}",
            t.name or "",
            t.phone or "",
            t.email or "",
            t.issue_category or "",
            t.description or "",
            t.status or "Open",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            t.resolved_at.strftime("%Y-%m-%d %H:%M") if t.resolved_at else "",
            t.resolution_note or "",
            t.resolved_by or "",
            t.customer_feedback or "",
            t.feedback_at.strftime("%Y-%m-%d %H:%M") if t.feedback_at else "",
            t.reopen_reason or "",
            str(t.reopen_count or 0)
        ])
                f"SAV-2025-{t.id:03d}",
                t.name or "",
                t.phone or "",
                t.email or "",
                t.issue_category or "",
                t.description or "",
                t.status or "Open",
                t.created_at.strftime("%Y-%m-%d %H:%M")
                    if t.created_at else ""
            ])
        
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(
                            max_len, len(str(cell.value))
                        )
                except:
                    pass
            ws.column_dimensions[col_letter].width = (
                min(max_len + 2, 50)
            )
        
        wb.save(filepath)
        print(f"Excel saved: {len(tickets)} tickets")
        
        return FileResponse(
            path=filepath,
            media_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            filename="savomart_support_tickets.xlsx",
            headers={
                "Content-Disposition": (
                    "attachment; "
                    "filename=savomart_support_tickets.xlsx"
                ),
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Expose-Headers": (
                    "Content-Disposition"
                )
            }
        )
    except Exception as e:
        print(f"Excel error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Excel error: {str(e)}"
        )
